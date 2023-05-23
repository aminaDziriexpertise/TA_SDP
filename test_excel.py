import openpyxl 
import requests

TA_EXISTANT = "http://127.0.0.1:5000/TA_EXISTANT"
SDP_EXISTANT = "http://127.0.0.1:5000/SDP_EXISTANT"
TA_PROJET = "http://127.0.0.1:5000/TA_PROJET"
SDP_PROJET = "http://127.0.0.1:5000/SDP_PROJET"




filepath="test_model.xlsx"
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet_ranges = workbook.create_sheet('SDP',0)

sheet_ranges['A1'] = "Etages"
sheet_ranges['B1'] = "Destinations"
sheet_ranges['C1'] = "Surface existante avant travaux (A)"
sheet_ranges['C2'] =  SDP_EXISTANT
sheet_ranges['D1'] = "Surface creee (B)"
sheet_ranges['E1'] = "Surface creee par changement de destination"
sheet_ranges['F1'] = "Surface demolie reconstruite"
sheet_ranges['G1'] = "Surface supprimee (D)"
sheet_ranges['H1'] = "Surface supprimee par changement de destination"
sheet_ranges['I1'] = "Surface projet"
sheet_ranges['I2'] = SDP_PROJET
sheet_ranges['J1'] = "Surface RDV"

sheet_ranges = workbook.create_sheet('TA')

sheet_ranges['A1'] = "Etages"
sheet_ranges['B1'] = "TA  existant"
sheet_ranges['B2'] = TA_EXISTANT
sheet_ranges['C1'] = "TA creee"
sheet_ranges['D1'] = "TA Demolie Reconstruite"
sheet_ranges['E1'] = "TA Supprimee"
sheet_ranges['F1'] = "TA projet"
sheet_ranges['F2'] = TA_PROJET
sheet_ranges['G1'] = "TA pour Stationnement"

std=workbook.get_sheet_by_name('Sheet')

workbook.remove_sheet(std)
# save workbook
workbook.save('input.xlsx')